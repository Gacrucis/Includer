import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import numpy as np
import os
import random
from data import man_names as men
from data import woman_names as women
from data import departamentos

NUM_STUDENTS = 1300
NUM_GRADES = 1000
NUM_DEBTS = 2000


class Personer():
    def __init__(self, wb, init_index=0, address_sheet_name='address.xlsx'):
        self.wb = wb
        self.id = init_index
        self.attrs = [
            "persona_id",
            "nombre",
            "identificacion",
            "celular",
            "direccion_fk",
            "sexo_fk",
            "tipo_identificacion_fk"]
        self.sheet = wb.active
        self.address_id = 0
        self.init_address_sheet(address_sheet_name)
        self.init_sheets()
        self.fill_people()

    def __del__(self):
        self.wb.save('Persona_profesor.xlsx')
        self.address_wb.save('Direccion.xlsx')

    def init_address_sheet(self, addr_name):
        try:
            self.address_wb = xl.load_workbook(addr_name)
        except FileNotFoundError:
            self.address_wb = xl.Workbook()
        self.address_sheet = self.address_wb.active
        self.address_attrs = [
            'direccion_id',
            'calle',
            'numero_a',
            'numero_b',
            'ciudad',
            'departamento'
        ]

    def init_sheets(self):
        max_col = len(self.attrs)

        for row in self.sheet.iter_rows(min_col=1, max_col=max_col):
            for header, cell in zip(self.attrs, row):
                cell.value = header
                cell.alignment = Alignment(horizontal='center')

        max_col = len(self.address_attrs)

        for row in self.address_sheet.iter_rows(min_col=1, max_col=max_col):
            for header, cell in zip(self.address_attrs, row):
                cell.value = header
                cell.alignment = Alignment(horizontal='center')

    def create_person(self, persona_id=None, nombre=None, identificacion=None, celular=None, direccion_fk=None,  sexo_fk=None, tipo_identificacion_fk=None):
        data = {}
        if not nombre:
            if random.random():
                nombre = random.sample(men, 1)[0]
                sexo_fk = 1
            else:
                nombre = random.sample(women, 1)[0]
                sexo_fk = 2
            if not sexo_fk:
                sexo_fk = random.randint(1, 2)
        else:
            first_name = nombre.split(' ')[0].capitalize()
            if first_name in men:
                sexo_fk = 1
            elif first_name in women:
                sexo_fk = 2
            else:
                if 'JATHINSON' in nombre:
                    sexo_fk = 1
                else:
                    sexo_fk = 3
        if not persona_id:
            self.id += 1
            persona_id = self.id
        if not identificacion:
            identificacion = random.randrange(1000000000, 9999999999)
        if not celular:
            celular = str(random.randint(
                300, 399)) + str(random.randint(100, 999)) + str(random.randint(1, 999)).zfill(3)

        direccion_fk = self.create_address(direccion_id=persona_id)

        if random.random() <= 0.9:
            tipo_identificacion_fk = 1
        else:
            tipo_identificacion_fk = 2

        args = list(locals().values())[1:-1]

        for attr, value in zip(self.attrs, args):
            data.setdefault(attr, value)

        return data

    def fill_people(self, sheet=None):
        """ write people attrs into a given worksheet """
        if not sheet:
            sheet = self.sheet
        professors_sheet = xl.load_workbook('Profesores.xlsx').active
        for cell in professors_sheet['A'][1:]:
            name = cell.value
            data = self.create_person(nombre=name)
            self.add(data, sheet)

    def create_address(self, direccion_id=None, calle=None, numero_a=None, numero_b=None, ciudad=None, departamento=None):
        data = {}
        if not departamento:
            if random.random() <= 0.85:
                departamento = 'Santander'
            else:
                departamento = random.sample(departamentos.keys(), 1)[0]

        if not ciudad:
            ciudades = departamentos[departamento]
            ciudad = random.sample(ciudades, 1)[0]
        else:
            if not ciudad in departamentos[departamento]:
                ciudad = random.sample(departamentos[departamento], 1)
        if not direccion_id:
            self.address_id += 1
            direccion_id = self.address_id
        else:
            if direccion_id <= self.address_id:
                direccion_id = self.address_id + 1
                self.address_id = direccion_id
        if not calle:
            calle = str(random.randint(1, 100))

        args = list(locals().values())[1:-1]

        for attr, value in zip(self.address_attrs, args):
            data.setdefault(attr, value)
        self.add(data, self.address_sheet, people=False)
        return data['direccion_id']

    def add(self, data, sheet, people=True):
        # target_row = 1
        letter = get_column_letter(len(data))
        row = sheet.max_row + 1
        for row in sheet[f"A{row}:{letter}{row}"]:
            if people:
                for attr, cell in zip(self.attrs, row):
                    cell.value = str(data[attr])
                    cell.alignment = Alignment(horizontal='center')
            else:
                for attr, cell in zip(self.address_attrs, row):
                    cell.value = str(data[attr])
                    cell.alignment = Alignment(horizontal='center')


def generate_grades(index, sheet, order, n=3, student_fk=None, asignatura_fk=None):

    for i in range(1, n+1):
        data = []
        grade = round(random.random() * 5, 2)
        data.extend([index + i, grade, student_fk, asignatura_fk])

        letter = get_column_letter(len(data))
        row_target = sheet.max_row + 1

        for row in sheet[f"A{row_target}:{letter}{row_target}"]:
            for attr, value, cell in zip(order, data, row):
                cell.value = str(value)
                cell.alignment = Alignment(horizontal='center')


def fill_grades(n, sheet):
    order = [
        'calificacion_id',
        'nota',
        'estudiante_fk',
        'asignatura_fk'
    ]
    set_headers(order, sheet)
    index = 0
    for i in range(n):
        generate_grades(index, sheet, order)
        index += 3


def generate_debts(sheet, description=None, tipo_deuda_fk=None, estudiante_fk=None):
    order = [
        'deuda_id',
        'cantidad',
        'descripcion',
        'tipo_deuda_fk',
        'estudiante_fk',
    ]
    max_col = len(order)

    for row in sheet.iter_rows(min_col=1, max_col=max_col):
        for header, cell in zip(order, row):
            cell.value = header
            cell.alignment = Alignment(horizontal='center')

    for i in range(1, NUM_DEBTS + 1):
        data = []
        value = random.randint(1, 100000)
        debt_type = random.randint(1, 4)
        data.extend([i, value, description, debt_type, estudiante_fk])

        letter = get_column_letter(len(data))
        row_target = sheet.max_row + 1
        for row in sheet[f"A{row_target}:{letter}{row_target}"]:
            for attr, value, cell in zip(order, data, row):
                cell.value = str(value)
                cell.alignment = Alignment(horizontal='center')


def generate_professors(sheet):
    order = [
        'profesor_id',
        'persona_fk',
        'escuela_fk'
    ]
    output_wb = xl.Workbook()
    output_sheet = output_wb.active
    max_col = len(order)
    for row in output_sheet.iter_rows(min_col=1, max_col=max_col):
        for header, cell in zip(order, row):
            cell.value = header
            cell.alignment = Alignment(horizontal='center')
    for index, cell in enumerate(sheet['A'][1:], 1):
        data = []
        if cell.value is None:
            continue
        persona_fk = cell.value
        if int(persona_fk) == 1561:
            escuela_fk = 18
        else:
            escuela_fk = random.randint(1, 29)

        data.extend([index, persona_fk, escuela_fk])
        add_bottom(data, output_sheet)

    output_wb.save('Professor.xlsx')


def fill_careers(sheet):
    output_wb = xl.Workbook()
    output_sheet = output_wb.active
    headers = ['administrador_carrera_id',
               'plan_carrera_numero_fk', 'estudiante_fk', 'carrera_fk']
    set_headers(headers, output_sheet)

    for index, cell in enumerate(sheet['A'][1:], 1):
        estudiante_id = cell.value
        carrera_id = random.randint(1, 30)
        plan_carrera_numero_fk = random.randint(1, 30)
        add_bottom([index, plan_carrera_numero_fk,
                    estudiante_id, carrera_id], output_sheet)

    output_wb.save('test\AdministradorCarrera.xlsx')


def fill_plan_signature(sheet):
    output_wb = xl.Workbook()
    output_sheet = output_wb.active
    headers = ['asignatura_plan_id',
               'asignatura_id', 'plan_estudios_id']
    set_headers(headers, output_sheet)
    signatures_id = [i for i in range(len(sheet['A'][1:]))]
    index = 1
    while signatures_id:
        random.shuffle(signatures_id)
        asignatura_plan_id = index
        asignatura_id = signatures_id[-1]
        signatures_id.pop()
        plan_estudios_id = random.randint(1, 30)
        add_bottom([asignatura_plan_id, asignatura_id,
                    plan_estudios_id], output_sheet)
        index += 1

    output_wb.save('test\AsignaturaPlan.xlsx')


def fill_deuda(sheet):
    output_wb = xl.Workbook()
    output_sheet = output_wb.active
    headers = ['deuda_id',
               'cantidad',
               'descripcion',
               'tipo_deuda_fk',
               'estudiante_fk']

    set_headers(headers, output_sheet)
    deuda_index = 0
    for index, cell in enumerate(sheet['A'][1:], 1):
        estudiante_fk = cell.value
        for i in range(1, 5):
            if random.random() >= 0.95:
                tipo_deuda = i
                cantidad = random.randint(100, 100000)
                deuda_index += 1
                add_bottom([deuda_index, cantidad, None,
                            tipo_deuda, estudiante_fk], output_sheet)

    output_wb.save('test\Deuda.xlsx')

def fill_history(sheet):
    output_wb = xl.Workbook()
    output_sheet = output_wb.active
    headers = ['historial_id',
               'estudiante_fk',
               'tipo_asignatura_fk',
               'estado_asignatura_fk']
    set_headers(headers, output_sheet)
    for index, cell in enumerate(sheet['A'][1:], 1):




def set_headers(headers, sheet):
    max_col = len(headers)
    for row in sheet.iter_rows(min_col=1, max_col=max_col):
        for header, cell in zip(headers, row):
            cell.value = header
            cell.alignment = Alignment(horizontal='center')


def add_bottom(data, sheet):
    letter = get_column_letter(len(data))
    row = sheet.max_row + 1
    ran = f"A{row}:{letter}{row}"
    for row in sheet[ran]:
        for value, cell in zip(data, row):
            cell.value = str(value)
            cell.alignment = Alignment(horizontal='center')


class Studener():
    def __init__(self, amount, wb):
        self.amount = amount
        self.id = 0
        self.attrs = [
            'estudiante_id',
            'codigo',
            'persona_fk',
            'historial_fk'
        ]
        self.wb = wb
        self.sheet = wb.active
        self.init_sheets()

    def __del__(self):
        self.wb.save('Estudiante.xlsx')
        # self.address_wb.save('Direccion.xlsx')

    def init_sheets(self):
        max_col = len(self.attrs)

        for row in self.sheet.iter_rows(min_col=1, max_col=max_col):
            for header, cell in zip(self.attrs, row):
                cell.value = header
                cell.alignment = Alignment(horizontal='center')

        self.persona_sheet = xl.load_workbook('Persona.xlsx').active

    def create_student(self, estudiante_id=None, codigo=None, persona_fk=None, historial_fk=None):
        data = []
        self.id += 1
        if estudiante_id:
            data.append(estudiante_id)
        else:
            data.append(self.id)
        if not codigo:
            codigo = '2' + str(random.randint(10, 21)) + \
                str(random.randint(0, 9999)).zfill(4)
            data.append(codigo)
        if not persona_fk:
            persona_fk = self.persona_sheet[f'A{self.id + 1}'].value
            data.append(persona_fk)
        data.append(None)  # historial_fk
        return data

    def fill_students(self, sheet=None):
        if not sheet:
            sheet = self.sheet

        for i in range(self.amount):
            data = self.create_student()
            self.add(data, sheet)

    def add(self, data, sheet):
        letter = get_column_letter(len(data))
        row = sheet.max_row + 1
        for row in sheet[f"A{row}:{letter}{row}"]:
            for value, cell in zip(data, row):
                cell.value = str(value)
                cell.alignment = Alignment(horizontal='center')


def main():
    try:
        wb = xl.load_workbook('files\Estudiante.xlsx')
    except FileNotFoundError:
        wb = xl.Workbook()
    ws = wb.active
    # pepe = Personer(wb, init_index=1300)
    # fill_grades(NUM_GRADES)
    fill_deuda(ws)


if __name__ == "__main__":
    main()
