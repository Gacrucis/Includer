import os
import sys
import openpyxl
from openpyxl.styles import Font
import datetime

from AcademicObjects import Subject, Group
from IncludingObjects import Schedule
from AppUtils import Logger
import data

import pickle

GROUP_PATH_PREFIX = 'Group_'
SUBJECT_PATH_PREFIX = 'Subject__'
SHIFT_PATH_PREFIX = 'Shift_'
# BUILDING_PATH_PREFIX = 'Building_'

TODAY = datetime.datetime.today()
TIMESTAMP = TODAY.strftime("%d_%h_%Y-%H_%M_%S")

PATH_EXT = '.xlsx'

def main():

    codes = [20255, 22957, 22958, 22959]
    # codes = [22959]

    codes = set(data.codes)
    print(len(codes))

    subjects = {}

    # for code in codes:

    #     try:

    #         subject = Subject(code)
    #         subjects[code] = subject
        
    #     except Exception as e:
    #         continue

    # file_subj = open('asignaturas.obj', 'wb')
    # pickle.dump(subjects, file_subj)

    subjects = pickle.load(open('asignaturas.obj', 'rb'))

    group_labels = [
        'codigo',
        'cantidad_estudiantes',
        'capacidad',
        'codigo_asignatura',
        'edificio',
        'salon',
        'profesor'
    ]

    groups_info = []


    subject_labels = [
        'codigo',
        'numero_creditos',
        'nombre',
        'tipo_asignatura_fk',
    ]

    subjects_info = []


    shift_labels = [
        'hora_inicio',
        'hora_fin',
        'dia_semana',
        'grupo_codigo',
        'asignatura_codigo',
        'edificio',
        'salon'
    ]

    shifts_info = []

    all_info = [
        (GROUP_PATH_PREFIX, group_labels, groups_info),
        (SUBJECT_PATH_PREFIX, subject_labels, subjects_info),
        (SHIFT_PATH_PREFIX, shift_labels, shifts_info),
    ]


    for code in subjects:

        subject = subjects[code]
        subject_info = []

        subject_info.extend([
            subject.code,
            4,
            subject.name,
            1
        ])

        subjects_info.append(subject_info)

        for group in subjects[code].groups.values():

            if not group.rooms or not group.teachers:
                continue
            
            group_info = []

            raw_room = [place.strip() for place in group.rooms[0].split('-')]
            
            building = raw_room[0]
            room = raw_room[1]

            group_info.extend([
                group.code,
                group.student_quantity,
                group.capacity,
                group.subject.code,
                building,
                room,
                group.teachers[0]
            ])
            
            groups_info.append(group_info)

            for weekday, shift in group.schedule.items():

                shift_info = []

                if len(shift) < 2:
                    shift = (shift[0], shift[0]+1)

                shift_info.extend([
                    shift[0],
                    shift[1],
                    weekday,
                    group.code,
                    group.subject.code,
                    building,
                    room
                ])

                shifts_info.append(shift_info)
    
    for name, label_list, info in all_info:

        file_name = f'{name}{TIMESTAMP}{PATH_EXT}'

        to_excel(label_list, attr_list=info, path=file_name)


    # wb = openpyxl.Workbook()
    # ws = wb.active
    # ws.title = 'Groups'

    # row = 1

    # room_col = 1
    # code_col = 2
    # capacity_col = 3
    # teacher_col = 4
    # schedule_col = 5
    # subject_col = 6

    # ws.cell(row, room_col).value = 'SALON'
    # ws.cell(row, code_col).value = 'CODIGO'
    # ws.cell(row, capacity_col).value = 'CAPACIDAD'
    # ws.cell(row, teacher_col).value = 'PROFESOR'
    # ws.cell(row, schedule_col).value = 'HORARIO'
    # ws.cell(row, subject_col).value = 'ASIGNATURA'

    # row += 1

    # for code in subjects:

    #     for group in subjects[code].groups.values():

    #         if not group.teachers:
    #             continue

    #         ws.cell(row, room_col).value = group.rooms[0] if len(group.rooms) >= 1 else 'NULL'
    #         ws.cell(row, code_col).value = group.code
    #         ws.cell(row, capacity_col).value = group.capacity
    #         ws.cell(row, teacher_col).value = group.teachers[0] if len(group.teachers) >= 1 else 'NULL'
    #         ws.cell(row, schedule_col).value = str(group.schedule)
    #         ws.cell(row, subject_col).value = group.subject.code

    #         row += 1

    # wb.save(f'{EXCEL_PATH_PREFIX}{timestamp}.xlsx')


def to_excel(label_list, attr_list, path='output.xlsx', sheetname='Info', check_integrity=True, start_row=1):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheetname

    column_names = list(enumerate(label_list, 1))

    row = start_row

    for name in column_names:

        ws.cell(row, name[0]).value = name[1]
    
    row += 1

    if check_integrity:

        processed_attributes = []

        for attributes in attr_list:

            is_complete = True

            for attribute in attributes:

                if attribute is None:

                    is_complete = False
            
            if is_complete:
                processed_attributes.append(attributes)
    
    else:
        processed_attributes = attr_list


    for attributes in processed_attributes:

        for i, attribute in enumerate(attributes):

            cell = ws.cell(row, i + 1)

            cell.value = attribute
            cell.font = Font(name='Consolas', size=10)
        
        row += 1
    
    wb.save(path)


if __name__ == "__main__":
    main()
