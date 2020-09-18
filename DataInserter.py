import openpyxl as xl
import psycopg2 as pc

import random


class DataBase():
    def __init__(self):
        try:
            self.db = pc.connect(
                dbname='bkfqvmiu',
                user='bkfqvmiu',
                password='kbc0-qYKx_NKBooqfoaf_xSoFYTyfaEb',
                host='lallah.db.elephantsql.com'
            )
        except Exception as ex:
            print(f"[ERROR] Algo fue mal. - {ex}")
        self.cursor = self.db.cursor()

        self.weekdays = [
            'LUNES',
            'MARTES',
            'MIERCOLES',
            'JUEVES',
            'VIERNES',
            'SABADO',
            'DOMINGO'
        ]

        self.init_files()
        self.gestioner()

    def init_files(self):

        self.sheets = {}
        self.simple_sheets = [
            'Sexo',
            'TipoCarrera',
            'TipoIdentificacion',
            'TipoDeuda',
            'TipoAsignatura',
            'DiaSemana',
            'Direccion',
            'Edificio',
            'Escuelas',
            'EstadoAsignatura',
            'Facultad',
            'PlanEstudios',
            'Carreras',
            'Persona',
            'Professor',
            'Estudiante',
            'AdministradorCarrera',
            'AsignaturaPlan'
        ]

        # for sheet in self.simple_sheets:

            self.sheets[sheet] = xl.load_workbook(f'files\{sheet}.xlsx').active

        # print(self.sheets)

        # More complex tables

        # self.complex_sheets = [
        #     'Franja',
        #     'Grupo',
        #     'Asignatura',
        #     'Edificio'
        # ]

        # for sheet in self.complex_sheets:

        #     self.sheets[sheet] = xl.load_workbook(f'files/{sheet}.xlsx').active

        # # Complementary sheets

        # self.complem_sheets = [
        #     'ProfesoresSync'
        # ]

        # for sheet in self.complem_sheets:

        #     self.sheets[sheet] = xl.load_workbook(f'files/{sheet}.xlsx').active

    def __del__(self):
        # save file and database
        self.cursor.close()
        # self.db.commit()
        self.db.close()

    def gestioner(self):
        """ call functions """
        # self.add_sexo()
        # self.add_tipo_carrera()
        # self.add_tipo_identificacion()
        # self.add_tipo_deuda()
        # self.add_tipo_asignatura()
        # self.add_dia_semana()
        # self.add_direccion()
        # self.add_edificio()
        # self.add_escuela()
        # self.add_estado_asignatura()
        # self.add_facultad()
        # self.add_plan_estudios()
        # self.add_carreras()
        # self.add_persona()
        self.add_profesor()
        # self.add_estudiante()
        # self.add_administrador_carrera()
        self.add_asignatura_plan()
        # self.add_grades()
        # self.add_gradestudent()

        # self.add_rooms()
        # self.add_groups()
        # self.add_shifts()

    def add_sexo(self):
        structure = f"""
            INSERT INTO Sexo (sexo_id, nombre)
            VALUES (?, ?)
        """
        queries = self.add(structure, self.sheets['Sexo'])

        self.to_text(queries, 'SEXO.sql')

    def add_tipo_carrera(self):
        structure = f"""
            INSERT INTO TipoCarrera (tipo_carrera_id, nombre, presencial)
            VALUES (?, ?, ?)
        """
        queries = self.add(structure, self.sheets['TipoCarrera'])

        self.to_text(queries, 'TipoCarrera.sql')

    def add_tipo_identificacion(self):
        structure = f"""
            INSERT INTO TipoIdentificacion (tipo_identificacion_id, nombre)
            VALUES (?, "?")
        """
        queries = self.add(structure, self.sheets['TipoIdentificacion'])

        self.to_text(queries, 'TipoIdentificacion.sql')

    def add_tipo_deuda(self):
        structure = f"""
            INSERT INTO TipoDeuda (tipo_deuda_id, nombre, maximo)
            VALUES (?, "?", ?)
        """
        queries = self.add(structure, self.sheets['TipoDeuda'])

        self.to_text(queries, 'TipoDeuda.sql')

    def add_tipo_asignatura(self):
        structure = f"""
            INSERT INTO TipoAsignatura (tipo_asignatura_id, descripcion)
            VALUES (?, '?')
        """
        queries = self.add(structure, self.sheets['TipoAsignatura'])

        self.to_text(queries, 'TipoAsignatura.sql')

    def add_dia_semana(self):
        structure = f"""
            INSERT INTO DiaSemana (dia_semana_id, nombre)
            VALUES (?, ?)
        """
        queries = self.add(structure, self.sheets['DiaSemana'])

        self.to_text(queries, 'DiaSemana.sql')

    def add_direccion(self):
        structure = f"""
            INSERT INTO Direccion (direccion_id, calle, numero_a, numero_b, "ciudad", "departamento")
            VALUES (?, ?, ?, ?, ?, ?)
        """
        queries = self.add(structure, self.sheets['Direccion'])

        self.to_text(queries, 'Direccion.sql')

    def add_edificio(self):
        structure = f"""
            INSERT INTO Edificio (edificio_id, nombre)
            VALUES (?, ?)
        """
        queries = self.add(structure, self.sheets['Edificio'])

        self.to_text(queries, 'Edificio.sql')

    def add_escuela(self):
        structure = f"""
            INSERT INTO Escuela (escuela_id, nombre, facultad_fk, edificio_fk)
            VALUES (?, "?", ?, ?)
        """
        queries = self.add(structure, self.sheets['Escuelas'])

        self.to_text(queries, 'Escuela.sql')

    def add_estado_asignatura(self):
        structure = f"""
            INSERT INTO EstadoAsignatura (estado_asignatura_id, descripcion)
            VALUES (?, '?')
        """
        queries = self.add(structure, self.sheets['EstadoAsignatura'])

        self.to_text(queries, 'EstadoAsignatura.sql')

    def add_facultad(self):
        structure = f"""
            INSERT INTO Facultad (facultad_id, facultad_nombre)
            VALUES (?, "?")
        """
        queries = self.add(structure, self.sheets['Facultad'])

        self.to_text(queries, 'Facultad.sql')

    def add_plan_estudios(self):
        structure = f"""
            INSERT INTO PlanEstudios (plan_estudios_id, numero_plan, carrera_fk)
            VALUES (?, ?, ?)
        """
        queries = self.add(structure, self.sheets['PlanEstudio'])

        self.to_text(queries, 'PlanEstudios.sql')

    def add_carreras(self):
        structure = f"""
            INSERT INTO Carrera (carrera_id, nombre, plan_actual, cantidad_semestres, escuela_fk, tipo_carrera_fk)
            VALUES (?, "?", ?, ?, ?, ?)
        """
        queries = self.add(structure, self.sheets['Carrera'])

        self.to_text(queries, 'Carrera.sql')

    def add_persona(self):
        structure = f"""
            INSERT INTO Persona (persona_id, "nombre", identificacion, celular, direccion_fk, sexo_fk, tipo_identificacion_fk)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """
        queries = self.add(structure, self.sheets['Persona'])

        self.to_text(queries, 'sql\Persona.sql')
    
    def add_profesor(self):
        structure = f"""
            INSERT INTO Profesor (profesor_id, persona_fk, escuela_fk)
            VALUES (?, ?, ?)
        """
        queries = self.add(structure, self.sheets['Professor'])

        self.to_text(queries, 'sql\Profesor.sql')

    def add_estudiante(self):
        structure = f"""
            INSERT INTO Estudiante (estudiante_id, codigo, persona_fk)
            VALUES (?, ?, ?)
        """
        queries = self.add(structure, self.sheets['Estudiante'])

        self.to_text(queries, 'sql\Estudiante.sql')

    def add_administrador_carrera(self):
        structure = f"""
            INSERT INTO AdministradorCarrera (administrador_carrera_id, plan_carrera_numero_fk, estudiante_fk, carrera_fk)
            VALUES (?, ?, ?, ?)
        """
        queries = self.add(structure, self.sheets['AdministradorCarrera'])

        self.to_text(queries, 'sql\AdministradorCarrera.sql')

    def add_asignatura_plan(self):
        structure = f"""
            INSERT INTO AsignaturaPlan (asignatura_plan_id, asignatura_fk, plan_estudios_fk)
            VALUES (?, ?, ?)
        """
        queries = self.add(structure, self.sheets['AsignaturaPlan'])

        self.to_text(queries, 'sql\AsignaturaPlan.sql') 

    def add(self, structure, sheet):
        """ structure is something like:
            INSERT INTO table_name (some attrs)
            VALUES (? ,? , ...)
        """
        ws = sheet
        max_col = ws.max_column + 1
        max_row = ws.max_row + 1
        queries = []
        for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            attrs = []
            for cell in row:
                if cell.value is None:
                    continue
                attrs.append(cell.value)
            if not attrs:
                continue
            query = structure
            for attr in attrs:
                query = query.replace('?', str(attr), 1)
            if not ';' in query:
                query += ';'
            queries.append(query)

        return queries

    def add_fromlist(self, structure, nestedlist):

        queries = []

        for elements in nestedlist:

            query = structure

            for attr in elements:

                query = query.replace('?', str(attr), 1)

            if not ';' in query:
                query += ';'

            queries.append(query)

        # for index, query in enumerate(queries):
        #     print(f"[{index}] -> {query}")

        return queries

    def add_rooms(self):

        structure = f"""
            INSERT INTO Salon (salon_id, codigo, capacidad, edificio_fk)
            VALUES (?, ?, ?, ?)
        """

        groups_sheet = self.sheets['Grupo']
        buildings_sheet = self.sheets['Edificio']

        self.buildings = {}

        for row in buildings_sheet.iter_rows(min_row=2):

            building_id = row[0].value

            if building_id is None:
                continue

            building_name = (row[1].value).upper()

            self.buildings.setdefault(building_name, int(building_id))

        self.room_queries = []
        self.room_sync = {}

        current_id = 1

        unique_ids = set()

        for row in groups_sheet.iter_rows(min_row=2):

            room_code = str(row[5].value).split(' ')[1].strip()
            room_capacity = int(row[2].value)

            try:

                room_building = self.buildings[str(row[4].value).upper()]

            except KeyError as e:

                room_building = random.randint(1, 20)

            room_uniqueid = f'{room_building}{room_code}'

            if room_uniqueid not in unique_ids:
                self.room_queries.append([
                    current_id,
                    room_code,
                    room_capacity,
                    room_building
                ])

            else:
                unique_ids.add(room_uniqueid)

            self.room_sync[room_uniqueid] = current_id

            current_id += 1

        queries = self.add_fromlist(structure, self.room_queries)

        self.to_text(queries, 'Salon.sql')

    def add_groups(self):

        structure = f"""
            INSERT INTO Grupo (grupo_id, codigo, cantidad_estudiantes, capacidad, codigo_asignatura, profesor_fk)
            VALUES (?, ?, ?, ?, ?, ?)
        """
        teacher_sync = self.sheets['ProfesoresSync']
        group_sheet = self.sheets['Grupo']

        self.teacher_sync = {}

        for row in teacher_sync.iter_rows(min_row=2):

            teacher_name = str(row[3].value).strip().upper()
            teacher_id = row[0].value

            self.teacher_sync.setdefault(teacher_name, teacher_id)

        self.group_queries = []

        self.shift_sync = {}

        current_id = 1

        for row in group_sheet.iter_rows(min_row=2):

            group_code = row[0].value
            group_quantity = row[1].value
            group_capacity = row[2].value
            group_subject = row[3].value
            group_teacher = str(row[6].value).strip().upper()

            try:
                group_teacher_fk = self.teacher_sync[group_teacher]

            except KeyError:
                group_teacher_fk = random.randint(1, 200)
                print('[INFO] Profesor no encontrado, usado uno aleatorio')

            self.shift_sync.setdefault(
                f'{group_subject}{group_code}', current_id)

            self.group_queries.append([
                current_id,
                group_code,
                group_quantity,
                group_capacity,
                group_subject,
                group_teacher_fk
            ])

            current_id += 1

        queries = self.add_fromlist(structure, self.group_queries)

        self.to_text(queries, 'Grupo.sql')

    def add_shifts(self):

        structure = f"""
            INSERT INTO Franja (franja_id, hora_inicio, hora_fin, salon_fk, grupo_fk, dia_semana_fk)
            VALUES (?, ?, ?, ?, ?, ?)
        """

        shifts_sheet = self.sheets['Franja']

        self.shift_queries = []

        current_id = 1

        for row in shifts_sheet.iter_rows(min_row=2):

            raw_building = str(row[5].value).upper().strip()

            try:
                building_num = self.buildings[raw_building]

            except KeyError:
                continue

            room_code = str(row[6].value).split(' ')[1].strip()

            room_uniqueid = f'{building_num}{room_code}'

            try:
                building_fk = self.room_sync[room_uniqueid]

            except KeyError:
                continue

            start_hour = int(row[0].value)
            end_hour = int(row[1].value)
            weekday_num = self.weekdays.index(
                str(row[2].value).strip().upper()) + 1

            group_code = row[3].value
            subject_code = row[4].value

            group_identifier = f'{subject_code}{group_code}'
            group_fk = self.shift_sync.get(
                group_identifier, random.randint(1, 200))

            self.shift_queries.append([
                current_id,
                start_hour,
                end_hour,
                building_fk,
                group_fk,
                weekday_num
            ])

            current_id += 1

            # TBA

            queries = self.add_fromlist(structure, self.shift_queries)

            self.to_text(queries, 'Franja.sql')

    def add_grades(self):

        structure = f"""
            INSERT INTO CalificacionAsignatura (calificacion_id, nota, estudiante_fk, asignatura_fk)
            VALUES (?, ?, ?, ?)
        """

        student_range = 1300
        subject_range = 247

        self.student_sync = {sid: [] for sid in range(1, student_range + 1)}

        self.grades_queries = []

        current_id = 1

        for sid in range(1, student_range + 1):

            for n in range(3):

                grade_unit = random.choice([0, 1, 2, 2, 3, 3, 3, 3, 4, 4, 4])
                grade_decimal = random.randint(0, 10)/10

                grade = grade_unit + grade_decimal

                rand_subject = random.choice(list(range(1, subject_range)))

                self.student_sync[sid].append(rand_subject)

                self.grades_queries.append([
                    current_id,
                    grade,
                    sid,
                    rand_subject
                ])

                current_id += 1

        queries = self.add_fromlist(structure, self.grades_queries)

        self.to_text(queries, 'CalificacionAsignatura.sql')

    def add_gradestudent(self):

        structure = f"""
            INSERT INTO AsignaturaEstudiante (asignatura_estudiante_id, estudiante_fk, asignatura_fk)
            VALUES (?, ?, ?)
        """

        current_id = 1

        self.gradestudent_queries = []

        for student, subjects in self.student_sync.items():

            for subject in subjects:

                self.gradestudent_queries.append([
                    current_id,
                    student,
                    subject
                ])

                current_id += 1

        queries = self.add_fromlist(structure, self.gradestudent_queries)

        self.to_text(queries, 'AsignaturaEstudiante.sql')

    def to_text(self, queries, path='out.sql'):

        with open(path, 'w') as f:

            for query in queries:

                f.write(str(query) + '\n')


def main():
    db = DataBase()


if __name__ == "__main__":
    main()
