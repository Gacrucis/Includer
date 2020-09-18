import os
import sys
import openpyxl
import datetime

from AcademicObjects import Subject, Group
from IncludingObjects import Schedule
from AppUtils import Logger

EXCEL_PATH_PREFIX = 'Groups_'


def main():

    codes = [20255, 22957, 22958, 22959]
    # codes = [22959]
    subjects = {}

    for code in codes:
        subject = Subject(code)

        subjects[code] = subject

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Groups'

    row = 1

    room_col = 1
    code_col = 2
    capacity_col = 3
    teacher_col = 4
    schedule_col = 5
    subject_col = 6

    ws.cell(row, room_col).value = 'SALON'
    ws.cell(row, code_col).value = 'CODIGO'
    ws.cell(row, capacity_col).value = 'CAPACIDAD'
    ws.cell(row, teacher_col).value = 'PROFESOR'
    ws.cell(row, schedule_col).value = 'HORARIO'
    ws.cell(row, subject_col).value = 'ASIGNATURA'

    row += 1

    for code in subjects:

        for group in subjects[code].groups.values():

            ws.cell(row, room_col).value = group.rooms[0] if len(group.rooms) >= 1 else 'NULL'
            ws.cell(row, code_col).value = group.code
            ws.cell(row, capacity_col).value = group.capacity
            ws.cell(row, teacher_col).value = group.teachers[0] if len(group.teachers) >= 1 else 'NULL'
            ws.cell(row, schedule_col).value = str(group.schedule)
            ws.cell(row, subject_col).value = group.subject.code

            row += 1

    today = datetime.datetime.today()
    timestamp = today.strftime("%d_%h_%Y-%H_%M_%S")

    wb.save(f'{EXCEL_PATH_PREFIX}{timestamp}.xlsx')


if __name__ == "__main__":
    main()
